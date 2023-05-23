import locale,openpyxl
from tqdm import tqdm
from openpyxl import load_workbook
import time
import os,xlwt
import pandas as pd

'''
读取Excel文件，输出指定的Sheet中指定的行和列
'''

def process_Excel(txt_path):
    #加载一个已存在的txt文件
    txt = load_workbook(txt_path)
    ws = txt.active
    lists = []
    for list in ws.iter_rows():
        lists.append(list)

    #print(lists)
    #print(lists[0])
    #print(lists[0][0].value)
    print(lists[1][1].value)
    return

'''
根据要求批量操作Excel文件
'''
def updata_excel(excel_path):
    filename = os.listdir(excel_path)
    for i in range(len(filename)):
        df = pd.read_excel(excel_path + filename[i])
        # df = df.drop(['content_type_id'],axis=1)
        df = df.drop(df[(df['id'] < 40)].index)

        df.to_excel(excel_path + filename[i])
        print(filename[i] + '********文件处理成功')

'''
实现指定的功能，并加入进度条

'''
def index():
    process_Excel(txt_path)
    for i in tqdm(range(1),desc='excel生成中',ncols=60):
        time.sleep(0.1)

if __name__ == '__main__':
    txt_path = ('/home/synsense/下载/1.xlsx')
    updata_excel(txt_path)